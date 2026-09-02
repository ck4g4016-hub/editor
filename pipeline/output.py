# -*- coding: utf-8 -*-
"""產生兩個輸出檔。規格依據見 docs/output-spec.md。

外網給 RPA 讀，用門牌到地政系統查地建號、下載謄本。
內網這一份**不是**戶政系統的上傳檔，是餵給內網腳本 2 的中繼檔 ——
腳本 2 讀它，自己轉出 upload1.xls／upload2.xls 再上傳。這點一開始弄錯過。

內網腳本 2（產製地址清冊並上傳至戶政系統）的兩行決定了格式：

    If LCase(fso.GetExtensionName(objFile.Name)) = "xlsx" Then   '第 234 行
    District = firstRow.RawData("行政區")                          '第 126 行

第一行：它在資料夾裡**只找 .xlsx**，csv、xls 一律看不到，找不到就跳
「資料夾內找不到 Excel 檔案」然後中止。所以這一份必須是 xlsx。

第二行：欄位是**用名稱取值**不是用位置，所以標題列一定要有、名稱要一字不差，
但多給一個「姓名」欄不會干擾它（腳本 2 不會去讀那一欄）。

還有一個坑：`GetTargetExcel` 撿到第一個 xlsx 就 `Exit Function`，
所以那個資料夾裡**只能有這一個 xlsx**。外網那份也是 xlsx，兩個混在一起
會撿到哪一個看檔案系統的順序，不是檔名順序 —— 所以兩份不要放同一個資料夾。
"""

import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 外網欄位。E~H 由 RPA 回寫，我們一定要留空。
# 尤其是 H（備註）—— 腳本每一列開頭會讀它，只要有內容整列就直接跳過，
# 不查、不回寫，畫面上也不會有任何提示。
OUTER_HEADERS = ["行政區", "門牌", "申請案號或事由", "身分證字號",
                 "段名(或代碼)", "地號", "建號", "備註", "姓名"]

# RPA 只讀到 H，排序也只到 D，所以新增的姓名放 I 欄不會影響它
OUTER_RPA_COLUMNS = 8

# 內網中繼檔的欄位。前四個是腳本 0 原本產的（它的第 56~59 行），
# 名稱必須一字不差，腳本 2 靠名稱取值。姓名是我們多加的，只給人核對用。
INNER_HEADERS = ["序號", "行政區", "所有權人IDN", "完整地址", "姓名"]

# 「完整地址」照腳本 0 的寫法要含縣市與行政區 —— 它的第 130 行
# `district = Mid(address, 4, 3)` 是從地址的第 4~6 個字取出行政區，
# 反推回去第 1~3 個字就是縣市。外網那份則相反，只留路街門牌。
INNER_CITY = "新北市"

_HEADER_FILL = PatternFill("solid", fgColor="EFE6D0")
_RESERVED_FILL = PatternFill("solid", fgColor="F5F5F5")


def roc_date(when=None):
    """民國年月日，例如 1150827。"""
    when = when or datetime.date.today()
    return "%d%02d%02d" % (when.year - 1911, when.month, when.day)


def outer_path(folder, when=None):
    return os.path.join(folder, "RPA-查調謄本清冊_%s.xlsx" % roc_date(when))


def inner_path(folder, serial=1, when=None):
    r"""內網檔名只能用英數 —— 系統的限制。

    腳本 2 的 `GetTargetExcel` 會用 `^(\d+)` 抓檔名開頭的數字當案號，
    要 9 位以上才算數。我們的檔名開頭是 HH，抓不到，案號會是空字串 ——
    腳本 0 自己產的「跨機關通報_….xlsx」開頭是中文，一樣抓不到，
    所以空的案號本來就是正常情況，不用為了這個去遷就檔名。
    """
    return os.path.join(folder, "HH%s_%02d.xlsx" % (roc_date(when), serial))


def write_outer(records, path):
    """外網 RPA 查調謄本清冊（xlsx）。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "查調清冊"

    sheet.append(OUTER_HEADERS)
    for index, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL if index <= OUTER_RPA_COLUMNS else _RESERVED_FILL
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        sheet.append([
            record.get("district", ""),
            record.get("address", ""),
            record.get("doc_number", ""),
            record.get("id_number", ""),
            "", "", "", "",          # 段名、地號、建號、備註 —— 留給 RPA
            record.get("name", ""),
        ])

    for column, width in zip("ABCDEFGHI", (10, 34, 16, 14, 14, 12, 12, 24, 12)):
        sheet.column_dimensions[column].width = width

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return path


def inner_address(record):
    """內網要的「完整地址」：縣市 + 行政區 + 門牌。

    我們平常把行政區跟門牌分開存，外網那份也要分開；
    但腳本 0 產的中繼檔是把三段黏成一串的，腳本 2 沿用那個寫法。
    """
    return INNER_CITY + record.get("district", "") + record.get("address", "")


def write_inner(records, path):
    """內網中繼檔（xlsx）—— 給腳本 2 讀，不是上傳檔。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "工作表1"

    sheet.append(INNER_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for serial, record in enumerate(records, start=1):
        sheet.append([
            serial,
            record.get("district", ""),
            record.get("id_number", ""),
            inner_address(record),
            record.get("name", ""),
        ])

    for column, width in zip("ABCDE", (8, 10, 14, 40, 12)):
        sheet.column_dimensions[column].width = width

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return path


# 內網系統一次匯入的上限
INNER_BATCH_LIMIT = 750


def write_all(records, folder, when=None):
    """兩個檔一起產。內網超過匯入上限就自動分批。"""
    written = [write_outer(records, outer_path(folder, when))]
    batches = [records[i:i + INNER_BATCH_LIMIT]
               for i in range(0, max(len(records), 1), INNER_BATCH_LIMIT)] or [[]]
    for serial, batch in enumerate(batches, start=1):
        written.append(write_inner(batch, inner_path(folder, serial, when)))
    return written
